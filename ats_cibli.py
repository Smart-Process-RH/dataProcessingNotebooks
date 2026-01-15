#!/usr/bin/env python3
"""
SCRIPT UNIQUE - Analyse Cabine Cibli - Mise à Jour Complète
=============================================================

Fusionne TOUS les processus:
1. Mise à jour des données depuis l'API
2. Filtrage par source (cabine cibli job)
3. Génération des statistiques
4. Export Excel

Date: 14 janvier 2026
Version: 3.0
"""

import requests
import pandas as pd
import numpy as np
import json
import warnings
import os
from datetime import datetime, timedelta
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

warnings.filterwarnings('ignore')

# ============================================================================
# ⚙️ CONFIGURATION
# ============================================================================

print("\n" + "="*80)
print("📊 ANALYSE CABINE CIBLI - SCRIPT UNIQUE v3.0")
print("="*80)

# Configuration - Modifier ces variables selon vos besoins
SOURCE_FILTER = "cabine cibli job"  # ← SOURCE À ANALYSER
DATE_START = "2025-09-11"           # ← DATE DE DÉBUT (YYYY-MM-DD)
DATE_END = "2026-01-14"             # ← DATE DE FIN (YYYY-MM-DD)
CLIENT_FILTER = "all"               # ← CHOIX CLIENT: "all" pour tous, ou nom spécifique
                                    # Clients disponibles (voir ci-dessous)
TOP_N_CLIENTS = 15                  # ← Nombre de top clients à afficher
TOP_N_CAMPAIGNS = 15                # ← Nombre de top campagnes à afficher

# Configuration API
API_URL = "https://api.smart-process-rh.com/v1"
API_KEY = "9TTaz70w8biMjvJ9Q5eIHZwVlQRNmjqAqiNzyGjfeI1S4nubpkSAL1h87FoNrlMv"

HEADERS = {
    "x-api-key": API_KEY,
    "Content-Type": "application/json"
}

# Créer les dossiers nécessaires
os.makedirs("stats/applications", exist_ok=True)
os.makedirs("stats/campaigns", exist_ok=True)
os.makedirs("exports", exist_ok=True)
os.makedirs("backups", exist_ok=True)

# Afficher la configuration
print("\n✅ Configuration chargée:")
print(f"   Source filtrée: {SOURCE_FILTER}")
print(f"   Période: {DATE_START} à {DATE_END}")
print(f"   Filtre client: {CLIENT_FILTER}")
print(f"   Top clients: {TOP_N_CLIENTS}")
print(f"   Top campagnes: {TOP_N_CAMPAIGNS}")

# ============================================================================
# 📥 ÉTAPE 1 : RÉCUPÉRATION DES DONNÉES DEPUIS L'API
# ============================================================================

print("\n" + "="*80)
print("📥 ÉTAPE 1: RÉCUPÉRATION DES DONNÉES DEPUIS L'API")
print("="*80)

def fetch_data(endpoint, description):
    """Récupérer les données depuis l'API"""
    try:
        print(f"\n   {description}...")
        response = requests.get(f"{API_URL}{endpoint}", headers=HEADERS, timeout=60)
        response.raise_for_status()
        data = response.json()

        if isinstance(data, dict) and 'data' in data:
            df = pd.DataFrame(data['data'])
        elif isinstance(data, list):
            df = pd.DataFrame(data)
        else:
            df = pd.DataFrame(data)

        print(f"   ✅ {len(df)} enregistrements récupérés")
        return df
    except Exception as e:
        print(f"   ❌ Erreur: {e}")
        return None

# Récupérer les données
applications_df = fetch_data("/applications/all", "1️⃣ Candidatures")
campaigns_df = fetch_data("/debug/campaigns", "2️⃣ Campagnes")
groups_df = fetch_data("/debug/groups", "3️⃣ Groupes")
brands_df = fetch_data("/debug/brands", "4️⃣ Marques")
units_df = fetch_data("/debug/units", "5️⃣ Unités")

if applications_df is None:
    print("\n❌ Impossible de récupérer les données. Arrêt.")
    exit(1)

# ============================================================================
# 💾 ÉTAPE 2 : SAUVEGARDE DES FICHIERS RAW
# ============================================================================

print("\n" + "="*80)
print("💾 ÉTAPE 2: SAUVEGARDE DES FICHIERS RAW")
print("="*80)

print("\n   Sauvegarde des fichiers bruts...")
applications_df.to_csv('stats/applications/raw_applications.csv', index=False, encoding='utf-8')
print(f"   ✅ raw_applications.csv ({len(applications_df)} candidatures)")

campaigns_df.to_csv('stats/campaigns/raw_campaigns.csv', index=False, encoding='utf-8')
print(f"   ✅ raw_campaigns.csv ({len(campaigns_df)} campagnes)")

# Créer un backup
timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
backup_dir = Path("backups")
backup_dir.mkdir(exist_ok=True)

applications_df.to_csv(f'backups/applications_backup_{timestamp}.csv', index=False, encoding='utf-8')
print(f"   ✅ Backup créé: applications_backup_{timestamp}.csv")

# ============================================================================
# 🔍 ÉTAPE 3 : FILTRAGE ET ENRICHISSEMENT
# ============================================================================

print("\n" + "="*80)
print("🔍 ÉTAPE 3: FILTRAGE ET ENRICHISSEMENT")
print("="*80)

# Convertir les dates et supprimer la timezone
applications_df['created_at'] = pd.to_datetime(
    applications_df['created_at'],
    errors='coerce',
    utc=True
).dt.tz_localize(None)

# Filtrer par source
print(f"\n   Filtrage par source: '{SOURCE_FILTER}'...")
df_filtered = applications_df[applications_df['source'] == SOURCE_FILTER].copy()
print(f"   ✅ {len(df_filtered)} candidatures filtrées")

# Filtrer par dates
print(f"   Filtrage par dates: {DATE_START} à {DATE_END}...")
date_start_dt = pd.to_datetime(DATE_START)
date_end_dt = pd.to_datetime(DATE_END)
df_filtered = df_filtered[
    (df_filtered['created_at'] >= date_start_dt) &
    (df_filtered['created_at'] <= date_end_dt)
]
print(f"   ✅ {len(df_filtered)} candidatures dans la période")

# Sauvegarder les candidatures filtrées
df_filtered.to_csv('stats/applications/cabine_cibli_job_applications.csv', index=False, encoding='utf-8')
print(f"   ✅ cabine_cibli_job_applications.csv sauvegardé")

# Fonction pour extraire les IDs
def extract_id(val):
    if pd.isna(val):
        return None
    if isinstance(val, dict):
        return val.get('id')
    if isinstance(val, str):
        try:
            import ast
            parsed = ast.literal_eval(val)
            if isinstance(parsed, dict):
                return parsed.get('id')
        except:
            pass
    return val

# Extraire les IDs
print("\n   Extraction des IDs...")
df_filtered['applicant_id'] = df_filtered['applicant'].apply(extract_id)
df_filtered['campaign_id'] = df_filtered['campaign'].apply(extract_id)
df_filtered['group_id'] = df_filtered.get('group', pd.Series()).apply(extract_id)
df_filtered['brand_id'] = df_filtered.get('brand', pd.Series()).apply(extract_id)
df_filtered['unit_id'] = df_filtered.get('unit', pd.Series()).apply(extract_id)
print(f"   ✅ IDs extraits")

# Créer les mappings pour enrichissement
print("\n   Création des mappings...")
campaign_mapping = {}
if campaigns_df is not None and len(campaigns_df) > 0:
    if 'id' in campaigns_df.columns and 'title' in campaigns_df.columns:
        campaign_mapping = dict(zip(campaigns_df['id'], campaigns_df['title']))

group_mapping = {}
if groups_df is not None and len(groups_df) > 0:
    if 'id' in groups_df.columns and 'name' in groups_df.columns:
        group_mapping = dict(zip(groups_df['id'], groups_df['name']))

brand_mapping = {}
if brands_df is not None and len(brands_df) > 0:
    if 'id' in brands_df.columns and 'name' in brands_df.columns:
        brand_mapping = dict(zip(brands_df['id'], brands_df['name']))

unit_mapping = {}
if units_df is not None and len(units_df) > 0:
    if 'id' in units_df.columns and 'name' in units_df.columns:
        unit_mapping = dict(zip(units_df['id'], units_df['name']))

# Ajouter les noms réels
print("\n   Ajout des noms réels...")
df_filtered['campaign_name'] = df_filtered['campaign_id'].map(campaign_mapping).fillna('Unknown Campaign')

# Fusionner avec les données de campagne pour obtenir les organisations (group, brand, unit)
if campaigns_df is not None and len(campaigns_df) > 0:
    # Créer des mappings pour les organisations
    group_id_to_name = {}
    brand_id_to_name = {}
    unit_id_to_name = {}

    if groups_df is not None and len(groups_df) > 0 and 'id' in groups_df.columns and 'name' in groups_df.columns:
        group_id_to_name = dict(zip(groups_df['id'].astype(int), groups_df['name']))

    if brands_df is not None and len(brands_df) > 0 and 'id' in brands_df.columns and 'name' in brands_df.columns:
        brand_id_to_name = dict(zip(brands_df['id'].astype(int), brands_df['name']))

    if units_df is not None and len(units_df) > 0 and 'id' in units_df.columns and 'name' in units_df.columns:
        unit_id_to_name = dict(zip(units_df['id'].astype(int), units_df['name']))

    # Créer un mapping campaign_id -> (group_name, brand_name, unit_name)
    campaign_org_mapping = {}
    for idx, row in campaigns_df.iterrows():
        camp_id = row.get('id')

        group_id = row.get('group_id')
        brand_id = row.get('brand_id')
        unit_id = row.get('unit_id')

        # Convertir les IDs en noms
        group_name = ''
        brand_name = ''
        unit_name = ''

        if pd.notna(group_id):
            try:
                group_name = group_id_to_name.get(int(group_id), '')
            except:
                pass

        if pd.notna(brand_id):
            try:
                brand_name = brand_id_to_name.get(int(brand_id), '')
            except:
                pass

        if pd.notna(unit_id):
            try:
                unit_name = unit_id_to_name.get(int(unit_id), '')
            except:
                pass

        # Créer le nom du client (organisation)
        client_parts = [g for g in [group_name, brand_name, unit_name] if g]
        client_name = ' > '.join(client_parts) if client_parts else 'Unknown Client'

        campaign_org_mapping[camp_id] = client_name

    df_filtered['client_name'] = df_filtered['campaign_id'].map(campaign_org_mapping).fillna('Unknown Client')
else:
    # Fallback si pas d'infos d'organisation
    df_filtered['client_name'] = df_filtered['campaign_name']

print(f"   ✅ Noms enrichis")

# ============================================================================
# 🏢 ÉTAPE 3.5 : APPLICATION DU FILTRE CLIENT
# ============================================================================

print("\n" + "="*80)
print("🏢 ÉTAPE 3.5: SÉLECTION ET FILTRAGE CLIENT")
print("="*80)

# Récupérer la liste unique des clients
unique_clients = sorted([c for c in df_filtered['client_name'].unique() if c != 'Unknown Client'])

# Créer/Mettre à jour le fichier de commentaire avec la liste des clients
clients_comment_file = 'CLIENTS_DISPONIBLES.txt'
with open(clients_comment_file, 'w', encoding='utf-8') as f:
    f.write("# ════════════════════════════════════════════════════════════════════════════════\n")
    f.write("# LISTE DES CLIENTS DISPONIBLES - MIS À JOUR AUTOMATIQUEMENT\n")
    f.write(f"# Généré le: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
    f.write("# ════════════════════════════════════════════════════════════════════════════════\n\n")
    f.write(f"# Total: {len(unique_clients)} clients disponibles\n\n")
    f.write("# 📋 LISTE DES CLIENTS:\n")
    f.write("# " + "─" * 76 + "\n")
    for idx, client in enumerate(unique_clients, 1):
        f.write(f"#   {idx:2d}. {client}\n")
    f.write("# " + "─" * 76 + "\n\n")
    f.write("# 🔧 UTILISATION DANS LE SCRIPT:\n")
    f.write("#   CLIENT_FILTER = \"all\"  # Pour analyser TOUS les clients\n")
    f.write("#   CLIENT_FILTER = \"NOM_EXACT_DU_CLIENT\"  # Pour un client spécifique\n\n")
    f.write(f"# ℹ️  Exemple (basé sur les clients actuels):\n")
    if unique_clients:
        f.write(f"#   CLIENT_FILTER = \"{unique_clients[0]}\"  # Pour le 1er client\n")

print(f"\n   ✅ {len(unique_clients)} clients trouvés")
print(f"   ✅ Liste sauvegardée dans: {clients_comment_file}")

# Afficher la liste des clients
print(f"\n   📋 Clients disponibles:\n")
for idx, client in enumerate(unique_clients, 1):
    client_count = len(df_filtered[df_filtered['client_name'] == client])
    print(f"      {idx:2d}. {client} ({client_count} candidatures)")

# Appliquer le filtre client
print(f"\n   Filtrage par client: '{CLIENT_FILTER}'...")
if CLIENT_FILTER.lower() != "all":
    if CLIENT_FILTER not in unique_clients:
        print(f"\n   ❌ ERREUR: Client '{CLIENT_FILTER}' non trouvé!")
        print(f"\n   Clients disponibles:")
        for client in unique_clients:
            print(f"      • {client}")
        print("\n   ❌ Arrêt du script.")
        exit(1)
    df_filtered = df_filtered[df_filtered['client_name'] == CLIENT_FILTER].copy()
    print(f"   ✅ Filtrage appliqué: {len(df_filtered)} candidatures pour '{CLIENT_FILTER}'")
else:
    print(f"   ✅ Tous les clients sélectionnés ({len(unique_clients)} clients, {len(df_filtered)} candidatures)")

# ============================================================================
# 📊 ÉTAPE 4 : STATISTIQUES GLOBALES
# ============================================================================

print("\n" + "="*80)
print("📊 ÉTAPE 4: STATISTIQUES GLOBALES")
print("="*80)

total_cvs = df_filtered['applicant_id'].nunique()
total_applications = len(df_filtered)
total_clients = df_filtered['campaign_id'].nunique()
total_campaigns = df_filtered['campaign_name'].nunique()

print(f"\n   ✓ Nombre total de CV faits: {total_cvs}")
print(f"   ✓ Nombre total de candidatures: {total_applications}")
print(f"   ✓ Nombre de clients: {total_clients}")
print(f"   ✓ Nombre de campagnes créées: {total_campaigns}")
print(f"\n   ✓ Période: {DATE_START} à {DATE_END}")
print(f"   ✓ Source: {SOURCE_FILTER}")

# ============================================================================
# 📈 ÉTAPE 5 : STATISTIQUES PAR PÉRIODE
# ============================================================================

print("\n" + "="*80)
print("📈 ÉTAPE 5: STATISTIQUES PAR PÉRIODE")
print("="*80)

# Convertir en période mensuelle
df_filtered['month'] = df_filtered['created_at'].dt.to_period('M')

# Stats globales
print("\n   1️⃣ Sauvegarde des stats globales...")
stats_total = pd.DataFrame({
    'metric': ['Total CVs', 'Total Candidatures', 'Total Clients'],
    'value': [total_cvs, total_applications, total_clients]
})
stats_total.to_csv('stats/cabin_stats_total_20260114.csv', index=False, encoding='utf-8')
print(f"   ✅ cabin_stats_total_20260114.csv")

# Stats par mois
print("\n   2️⃣ Calcul par mois...")
for month in df_filtered['month'].dropna().unique():
    df_month = df_filtered[df_filtered['month'] == month]

    if month == pd.Period('2025-12', 'M'):
        filename = 'stats/cabin_stats_december_20260114.csv'
        label = "Décembre 2025"
    elif month == pd.Period('2026-01', 'M'):
        filename = 'stats/cabin_stats_january_20260114.csv'
        label = "Janvier 2026"
    else:
        continue

    stats_month = pd.DataFrame({
        'metric': ['CVs', 'Candidatures', 'Clients'],
        'value': [
            df_month['applicant_id'].nunique(),
            len(df_month),
            df_month['campaign_id'].nunique()
        ]
    })
    stats_month.to_csv(filename, index=False, encoding='utf-8')
    print(f"   ✅ {label}: {len(df_month)} candidatures")

# ============================================================================
# 📅 ÉTAPE 6 : STATISTIQUES PAR MOIS (DÉTAILLÉES)
# ============================================================================

print("\n" + "="*80)
print("📅 STATISTIQUES PAR MOIS")
print("="*80)

print("\n   Détail par mois:\n")
for month in sorted(df_filtered['month'].dropna().unique()):
    df_month = df_filtered[df_filtered['month'] == month]
    month_str = str(month)
    month_cvs = df_month['applicant_id'].nunique()
    month_apps = len(df_month)
    month_clients = df_month['campaign_id'].nunique()

    print(f"   📅 {month_str}:")
    print(f"      • CVs: {month_cvs}")
    print(f"      • Candidatures: {month_apps}")
    print(f"      • Clients: {month_clients}")

# ============================================================================
# 🏢 ÉTAPE 7 : TOP CLIENTS
# ============================================================================

print("\n" + "="*80)
print("🏢 TOP CLIENTS")
print("="*80)

# Grouper par client_name pour compter les candidatures
client_stats = df_filtered['client_name'].value_counts().head(TOP_N_CLIENTS)

print(f"\n   Top {min(TOP_N_CLIENTS, len(client_stats))} clients:\n")
for idx, (client, count) in enumerate(client_stats.items(), 1):
    percentage = (count / total_applications) * 100
    print(f"   {idx:2d}. {client}: {count} candidatures ({percentage:.1f}%)")

# ============================================================================
# 📊 ÉTAPE 8 : RÉPARTITION PAR STATUT
# ============================================================================

print("\n" + "="*80)
print("📊 RÉPARTITION PAR STATUT")
print("="*80)

if 'status' in df_filtered.columns:
    status_dist = df_filtered['status'].value_counts()
    print()
    for status, count in status_dist.items():
        percentage = (count / total_applications) * 100
        print(f"   - {status}: {count} ({percentage:.1f}%)")

# ============================================================================
# 🎯 ÉTAPE 9 : TOP CAMPAGNES
# ============================================================================

print("\n" + "="*80)
print("🎯 TOP CAMPAGNES")
print("="*80)

# Grouper par campaign_name pour compter les candidatures
campaign_counts = df_filtered['campaign_name'].value_counts().head(TOP_N_CAMPAIGNS)

print(f"\n   Top {min(TOP_N_CAMPAIGNS, len(campaign_counts))} campagnes:\n")
for idx, (campaign, count) in enumerate(campaign_counts.items(), 1):
    percentage = (count / total_applications) * 100
    print(f"   {idx:2d}. {campaign}: {count} candidatures ({percentage:.1f}%)")

# ============================================================================
# 💾 ÉTAPE 10 : EXPORT EXCEL
# ============================================================================

print("\n" + "="*80)
print("💾 EXPORT EXCEL")
print("="*80)

try:
    wb = Workbook()
    wb.remove(wb.active)

    # Styles
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    title_font = Font(bold=True, size=14, color="4472C4")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # ===== FEUILLE 1: RÉSUMÉ =====
    ws = wb.create_sheet("📋 Résumé", 0)

    ws['A1'] = "📊 ANALYSE CANDIDATURES CABINE CIBLI"
    ws['A1'].font = title_font
    ws.merge_cells('A1:B1')

    ws['A2'] = f"Période: {DATE_START} à {DATE_END}"
    ws['A2'].font = Font(italic=True, size=10)
    ws.merge_cells('A2:B2')

    row = 4
    ws[f'A{row}'] = "Métrique"
    ws[f'B{row}'] = "Valeur"
    ws[f'A{row}'].fill = header_fill
    ws[f'B{row}'].fill = header_fill
    ws[f'A{row}'].font = header_font
    ws[f'B{row}'].font = header_font

    row += 1
    summary_metrics = [
        ("SOURCE", SOURCE_FILTER),
        ("Nombre total de CV faits", total_cvs),
        ("Nombre total de candidatures", total_applications),
        ("Nombre de clients", total_clients),
        ("Date de début", DATE_START),
        ("Date de fin", DATE_END),
        ("Date de génération", datetime.now().strftime('%Y-%m-%d %H:%M:%S')),
    ]

    for metric_name, metric_value in summary_metrics:
        ws[f'A{row}'] = metric_name
        ws[f'B{row}'] = metric_value
        ws[f'A{row}'].border = border
        ws[f'B{row}'].border = border
        row += 1

    # ===== FEUILLE 2: TOP CLIENTS =====
    ws = wb.create_sheet("🏢 Top Clients", 1)

    ws['A1'] = "Rang"
    ws['B1'] = "Client (Organisation)"
    ws['C1'] = "Candidatures"
    ws['D1'] = "%"
    for col in ['A', 'B', 'C', 'D']:
        ws[f'{col}1'].fill = header_fill
        ws[f'{col}1'].font = header_font

    for idx, (client, count) in enumerate(client_stats.items(), 1):
        percentage = (count / total_applications) * 100
        ws[f'A{idx+1}'] = idx
        ws[f'B{idx+1}'] = client
        ws[f'C{idx+1}'] = count
        ws[f'D{idx+1}'] = f"{percentage:.1f}%"
        for col in ['A', 'B', 'C', 'D']:
            ws[f'{col}{idx+1}'].border = border

    # ===== FEUILLE 3: TOP CAMPAGNES =====
    ws = wb.create_sheet("🎯 Top Campagnes", 2)

    ws['A1'] = "Rang"
    ws['B1'] = "Campagne (Titre)"
    ws['C1'] = "Candidatures"
    ws['D1'] = "%"
    for col in ['A', 'B', 'C', 'D']:
        ws[f'{col}1'].fill = header_fill
        ws[f'{col}1'].font = header_font

    for idx, (campaign, count) in enumerate(campaign_counts.items(), 1):
        percentage = (count / total_applications) * 100
        ws[f'A{idx+1}'] = idx
        ws[f'B{idx+1}'] = campaign
        ws[f'C{idx+1}'] = count
        ws[f'D{idx+1}'] = f"{percentage:.1f}%"
        for col in ['A', 'B', 'C', 'D']:
            ws[f'{col}{idx+1}'].border = border

    # ===== FEUILLE 4: STATUTS =====
    if 'status' in df_filtered.columns:
        ws = wb.create_sheet("📊 Statuts", 3)

        ws['A1'] = "Statut"
        ws['B1'] = "Nombre"
        ws['C1'] = "%"
        for col in ['A', 'B', 'C']:
            ws[f'{col}1'].fill = header_fill
            ws[f'{col}1'].font = header_font

        for idx, (status, count) in enumerate(status_dist.items(), 1):
            percentage = (count / total_applications) * 100
            ws[f'A{idx+1}'] = status
            ws[f'B{idx+1}'] = count
            ws[f'C{idx+1}'] = f"{percentage:.1f}%"
            for col in ['A', 'B', 'C']:
                ws[f'{col}{idx+1}'].border = border

    # Ajuster les largeurs
    for ws_name in wb.sheetnames:
        ws = wb[ws_name]
        for col in ['A', 'B', 'C', 'D']:
            ws.column_dimensions[col].width = 20

    # Sauvegarder
    excel_file = f"exports/cabine_cibli_analytics_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb.save(excel_file)

    print(f"\n   ✅ Fichier Excel généré: {excel_file}")

except Exception as e:
    print(f"\n   ❌ Erreur Excel: {e}")

# ============================================================================
# ✅ RÉSUMÉ FINAL
# ============================================================================

print("\n" + "="*80)
print("✅ ANALYSE COMPLÉTÉE")
print("="*80)

print(f"""
📊 RÉSUMÉ - CABINE CIBLI JOB

Source:                {SOURCE_FILTER}
Période:               {DATE_START} à {DATE_END}

CVs:                   {total_cvs}
Candidatures:          {total_applications}
Clients:               {total_clients}
Campagnes créées:      {total_campaigns}

📁 Fichiers générés:
   ✅ stats/applications/raw_applications.csv
   ✅ stats/applications/cabine_cibli_job_applications.csv
   ✅ stats/campaigns/raw_campaigns.csv
   ✅ stats/cabin_stats_total_20260114.csv
   ✅ stats/cabin_stats_december_20260114.csv
   ✅ stats/cabin_stats_january_20260114.csv
   ✅ exports/cabine_cibli_analytics_*.xlsx

📅 Date: 14 janvier 2026
""")

print("="*80 + "\n")

